from fastapi import FastAPI, UploadFile, Body, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse

import os
import shutil
import uuid
from io import StringIO, BytesIO
import csv
import pandas as pd
from datetime import datetime
from openpyxl.utils import get_column_letter

from database import init_db, get_connection
from storage import upload_to_minio, download_to_temp, object_exists, read_text
from ocr import extract_text
from classify import classify_text
from parsers import parse_invoice, parse_quote


# -----------------------------------------------------------------------------
# App & CORS
# -----------------------------------------------------------------------------
app = FastAPI()
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)
os.makedirs("tmp", exist_ok=True)
init_db()

# Migrazione: aggiunge le 3 colonne extra se mancano
def _ensure_invoices_extra_columns():
    conn = get_connection()
    try:
        cur = conn.cursor()
        cur.execute(
            """
            ALTER TABLE invoices
              ADD COLUMN IF NOT EXISTS supplier_vat_id   TEXT,
              ADD COLUMN IF NOT EXISTS supplier_tax_code TEXT,
              ADD COLUMN IF NOT EXISTS account_holder    TEXT;
            """
        )
        conn.commit()
    finally:
        conn.close()

_ensure_invoices_extra_columns()


# -----------------------------------------------------------------------------
# Upload & listing
# -----------------------------------------------------------------------------
@app.post("/documents/upload")
async def upload_document(file: UploadFile):
    if not file or not file.filename:
        raise HTTPException(status_code=400, detail="File mancante")

    tmp_basename = f"{uuid.uuid4()}_{os.path.basename(file.filename)}"
    tmp_path = os.path.join("tmp", tmp_basename)
    txt_path = None

    try:
        # 1) Salva temporaneamente
        with open(tmp_path, "wb") as buf:
            shutil.copyfileobj(file.file, buf)

        # 2) Carica originale su S3
        raw_key = f"tenant-1/raw/{tmp_basename}"
        storage_path = upload_to_minio(tmp_path, raw_key)

        # 3) Estrai testo + classifica
        text, used_ocr = extract_text(tmp_path)
        label, conf = classify_text(text)

        # 4) Carica testo su S3
        txt_path = os.path.join("tmp", tmp_basename + ".txt")
        with open(txt_path, "w", encoding="utf-8") as f:
            f.write(text)
        text_key = f"tenant-1/text/{tmp_basename}.txt"
        text_storage_path = upload_to_minio(txt_path, text_key)

        # 5) Registra su DB
        conn = get_connection()
        try:
            cur = conn.cursor()
            cur.execute(
                """
                INSERT INTO documents (filename, storage_path, doc_type, confidence, text_storage_path)
                VALUES (%s, %s, %s, %s, %s)
                """,
                (file.filename, storage_path, label, conf, text_storage_path),
            )
            conn.commit()
        finally:
            conn.close()

        return {
            "ok": True,
            "storage_path": storage_path,
            "doc_type": label,
            "confidence": conf,
            "text_storage_path": text_storage_path,
            "used_ocr": used_ocr,
        }

    finally:
        try:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
            if txt_path and os.path.exists(txt_path):
                os.remove(txt_path)
        except:
            pass


@app.get("/documents")
async def list_documents():
    conn = get_connection()
    try:
        cur = conn.cursor()
        cur.execute(
            """
            SELECT id, filename, doc_type, confidence, storage_path, uploaded_at
            FROM documents
            ORDER BY uploaded_at DESC
            """
        )
        rows = cur.fetchall()
        return [
            {
                "id": r[0],
                "filename": r[1],
                "doc_type": r[2],
                "confidence": float(r[3]) if r[3] is not None else None,
                "storage_path": r[4],
                "uploaded_at": r[5].isoformat(),
            }
            for r in rows
        ]
    finally:
        conn.close()


# -----------------------------------------------------------------------------
# Review tipo documento
# -----------------------------------------------------------------------------
@app.post("/documents/{doc_id}/review")
async def review_document(doc_id: int, new_type: str = Body(..., embed=True)):
    conn = get_connection()
    try:
        cur = conn.cursor()
        cur.execute(
            "UPDATE documents SET doc_type=%s, confidence=1.0 WHERE id=%s",
            (new_type, doc_id),
        )
        if cur.rowcount == 0:
            raise HTTPException(status_code=404, detail="Documento non trovato")
        conn.commit()
        return {"ok": True}
    finally:
        conn.close()


# -----------------------------------------------------------------------------
# Rielabora mancanti (OCR + classifica per vecchi record)
# -----------------------------------------------------------------------------
@app.post("/documents/reprocess_missing")
async def reprocess_missing():
    conn = get_connection()
    try:
        cur = conn.cursor()
        cur.execute(
            """
            SELECT id, filename, storage_path
            FROM documents
            WHERE doc_type IS NULL OR confidence IS NULL
            ORDER BY uploaded_at ASC
            """
        )
        rows = cur.fetchall()

        processed = 0
        skipped_missing = 0
        skipped_error = 0

        for doc_id, filename, storage_path in rows:
            local = None
            txt_path = None
            try:
                if not storage_path or not object_exists(storage_path):
                    skipped_missing += 1
                    continue

                local = download_to_temp(storage_path, "tmp")

                text, _ = extract_text(local)
                label, conf = classify_text(text)

                txt_name = os.path.basename(local) + ".txt"
                txt_path = os.path.join("tmp", txt_name)
                with open(txt_path, "w", encoding="utf-8") as f:
                    f.write(text)
                text_key = f"tenant-1/text/{txt_name}"
                text_storage_path = upload_to_minio(txt_path, text_key)

                cur.execute(
                    """
                    UPDATE documents
                    SET doc_type=%s, confidence=%s, text_storage_path=%s
                    WHERE id=%s
                    """,
                    (label, conf, text_storage_path, doc_id),
                )
                processed += 1

            except FileNotFoundError:
                skipped_missing += 1
            except Exception:
                skipped_error += 1
            finally:
                try:
                    if local and os.path.exists(local):
                        os.remove(local)
                    if txt_path and os.path.exists(txt_path):
                        os.remove(txt_path)
                except:
                    pass

        conn.commit()
        return {
            "ok": True,
            "processed": processed,
            "skipped_missing": skipped_missing,
            "skipped_error": skipped_error,
        }
    finally:
        conn.close()


# -----------------------------------------------------------------------------
# Reprocess strutturato (fatture/preventivi -> tabelle)
# -----------------------------------------------------------------------------
@app.post("/structured/reprocess")
async def reprocess_structured():
    conn = get_connection()
    try:
        cur = conn.cursor()
        cur.execute(
            """
            SELECT d.id, d.doc_type, d.text_storage_path, d.storage_path
            FROM documents d
            WHERE d.doc_type IN ('fattura','preventivo')
              AND NOT EXISTS (SELECT 1 FROM invoices i WHERE i.document_id = d.id)
              AND NOT EXISTS (SELECT 1 FROM quotes   q WHERE q.document_id = d.id)
            ORDER BY d.uploaded_at ASC
            """
        )
        rows = cur.fetchall()

        created_invoices = 0
        created_quotes = 0
        skipped_missing = 0
        skipped_error = 0

        for doc_id, doc_type, text_path, raw_path in rows:
            local = None
            try:
                # testo già estratto se presente
                if text_path and object_exists(text_path):
                    text = read_text(text_path)
                # altrimenti OCR al volo
                elif raw_path and object_exists(raw_path):
                    local = download_to_temp(raw_path, "tmp")
                    text, _ = extract_text(local)
                else:
                    skipped_missing += 1
                    continue

                if not text or not text.strip():
                    skipped_missing += 1
                    continue

                if doc_type == "fattura":
                    data = parse_invoice(text)
                    cur.execute(
                        """
                        INSERT INTO invoices
                        (document_id, supplier, supplier_vat_id, supplier_tax_code, account_holder,
                         number, date, net_amount, vat_amount, total_amount, currency, confidence)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                        ON CONFLICT (document_id) DO NOTHING
                        """,
                        (
                            doc_id,
                            data.get("supplier"),
                            data.get("supplier_vat_id"),
                            data.get("supplier_tax_code"),
                            data.get("account_holder"),
                            data.get("number"),
                            data.get("date"),
                            data.get("net_amount"),
                            data.get("vat_amount"),
                            data.get("total_amount"),
                            data.get("currency"),
                            data.get("confidence"),
                        ),
                    )
                    created_invoices += 1

                elif doc_type == "preventivo":
                    data = parse_quote(text)
                    cur.execute(
                        """
                        INSERT INTO quotes
                        (document_id, customer, valid_until, total_amount, items_json, confidence)
                        VALUES (%s,%s,%s,%s,%s,%s)
                        ON CONFLICT (document_id) DO NOTHING
                        """,
                        (
                            doc_id,
                            data.get("customer"),
                            data.get("valid_until"),
                            data.get("total_amount"),
                            data.get("items_json"),
                            data.get("confidence"),
                        ),
                    )
                    created_quotes += 1

            except FileNotFoundError:
                skipped_missing += 1
            except Exception:
                skipped_error += 1
            finally:
                try:
                    if local and os.path.exists(local):
                        os.remove(local)
                except:
                    pass

        conn.commit()
        return {
            "ok": True,
            "invoices": created_invoices,
            "quotes": created_quotes,
            "skipped_missing": skipped_missing,
            "skipped_error": skipped_error,
        }
    finally:
        conn.close()


# -----------------------------------------------------------------------------
# Liste
# -----------------------------------------------------------------------------
@app.get("/invoices")
async def list_invoices():
    conn = get_connection()
    try:
        cur = conn.cursor()
        cur.execute(
            """
            SELECT id, document_id,
                   supplier, supplier_vat_id, supplier_tax_code, account_holder,
                   number, date, net_amount, vat_amount, total_amount, currency, confidence, created_at
            FROM invoices
            ORDER BY created_at DESC
            """
        )
        rows = cur.fetchall()
        return [
            {
                "id": r[0],
                "document_id": r[1],
                "supplier": r[2],
                "supplier_vat_id": r[3],
                "supplier_tax_code": r[4],
                "account_holder": r[5],
                "number": r[6],
                "date": (r[7] and r[7].isoformat()),
                "net_amount": float(r[8]) if r[8] is not None else None,
                "vat_amount": float(r[9]) if r[9] is not None else None,
                "total_amount": float(r[10]) if r[10] is not None else None,
                "currency": r[11],
                "confidence": float(r[12]) if r[12] is not None else None,
                "created_at": r[13].isoformat(),
            }
            for r in rows
        ]
    finally:
        conn.close()


@app.get("/quotes")
async def list_quotes():
    conn = get_connection()
    try:
        cur = conn.cursor()
        cur.execute(
            """
            SELECT id, document_id, customer, valid_until, total_amount, items_json, confidence, created_at
            FROM quotes
            ORDER BY created_at DESC
            """
        )
        rows = cur.fetchall()
        return [
            {
                "id": r[0],
                "document_id": r[1],
                "customer": r[2],
                "valid_until": (r[3] and r[3].isoformat()),
                "total_amount": float(r[4]) if r[4] is not None else None,
                "items_json": r[5],
                "confidence": float(r[6]) if r[6] is not None else None,
                "created_at": r[7].isoformat(),
            }
            for r in rows
        ]
    finally:
        conn.close()


# -----------------------------------------------------------------------------
# Helpers export IT
# -----------------------------------------------------------------------------
def _fmt_date_it(val) -> str:
    if not val:
        return ""
    try:
        if isinstance(val, str):
            dt = datetime.fromisoformat(val[:10])
        else:
            dt = datetime(val.year, val.month, val.day)
        return dt.strftime("%d/%m/%Y")
    except Exception:
        return str(val)


def _fmt_num_it(x) -> str:
    if x is None:
        return ""
    try:
        s = f"{float(x):,.2f}"
        # 1,234.56 -> 1.234,56
        return s.replace(",", "§").replace(".", ",").replace("§", ".")
    except Exception:
        return str(x)


# -----------------------------------------------------------------------------
# Export (IT)
# -----------------------------------------------------------------------------
@app.get("/invoices/export")
async def export_invoices(format: str = "csv"):
    conn = get_connection()
    try:
        cur = conn.cursor()
        cur.execute(
            """
            SELECT supplier, supplier_vat_id, supplier_tax_code, account_holder,
                   number, date, net_amount, vat_amount, total_amount, currency, confidence
            FROM invoices
            ORDER BY created_at DESC
            """
        )
        rows = cur.fetchall()
    finally:
        conn.close()

    cols_excel = [
        "Fornitore", "P.IVA", "Cod. Fiscale", "Intestatario",
        "Numero", "Data", "Imponibile", "IVA", "Totale", "Valuta", "Confidenza"
    ]

    if format.lower() in ("csv", "csv_it", "csv-it"):
        sio = StringIO()
        w = csv.writer(sio, delimiter=";")
        w.writerow(cols_excel)
        for r in rows:
            w.writerow([
                r[0] or "",            # Fornitore
                r[1] or "",            # P.IVA
                r[2] or "",            # Cod. Fiscale
                r[3] or "",            # Intestatario
                r[4] or "",            # Numero
                _fmt_date_it(r[5]),    # Data
                _fmt_num_it(r[6]),     # Imponibile
                _fmt_num_it(r[7]),     # IVA
                _fmt_num_it(r[8]),     # Totale
                r[9] or "",            # Valuta
                _fmt_num_it(r[10]),    # Confidenza
            ])
        sio.seek(0)
        return StreamingResponse(
            iter([sio.getvalue()]),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": "attachment; filename=fatture.csv"},
        )

    # ---------- Excel ----------
    df = pd.DataFrame(
        rows,
        columns=[
            "supplier", "supplier_vat_id", "supplier_tax_code", "account_holder",
            "number", "date", "net_amount", "vat_amount", "total_amount", "currency", "confidence",
        ],
    )
    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    for c in ["net_amount", "vat_amount", "total_amount", "confidence"]:
        df[c] = pd.to_numeric(df[c], errors="coerce")

    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        sheet = "Fatture"
        df_ren = df.rename(columns={
            "supplier": "Fornitore",
            "supplier_vat_id": "P.IVA",
            "supplier_tax_code": "Cod. Fiscale",
            "account_holder": "Intestatario",
            "number": "Numero",
            "date": "Data",
            "net_amount": "Imponibile",
            "vat_amount": "IVA",
            "total_amount": "Totale",
            "currency": "Valuta",
            "confidence": "Confidenza",
        })[[
            "Fornitore","P.IVA","Cod. Fiscale","Intestatario",
            "Numero","Data","Imponibile","IVA","Totale","Valuta","Confidenza"
        ]]

        df_ren.to_excel(writer, index=False, sheet_name=sheet)
        ws = writer.sheets[sheet]

        widths = {
            "Fornitore": 28, "P.IVA": 14, "Cod. Fiscale": 16, "Intestatario": 20,
            "Numero": 16, "Data": 12, "Imponibile": 14, "IVA": 12, "Totale": 14,
            "Valuta": 10, "Confidenza": 12
        }
        for i, name in enumerate(df_ren.columns, start=1):
            ws.column_dimensions[get_column_letter(i)].width = widths.get(name, 16)

        # Data: forza datetime + formato dd/mm/yyyy
        date_col = list(df_ren.columns).index("Data") + 1
        for row in range(2, ws.max_row + 1):
            c = ws.cell(row=row, column=date_col)
            v = c.value
            if isinstance(v, str):
                try:
                    c.value = datetime.fromisoformat(v[:10])
                except Exception:
                    try:
                        parsed = pd.to_datetime(v, errors="coerce")
                        c.value = parsed.to_pydatetime() if pd.notna(parsed) else None
                    except Exception:
                        c.value = None
            c.number_format = "dd/mm/yyyy"

        # Numeri: formato italiano
        for name in ["Imponibile", "IVA", "Totale", "Confidenza"]:
            col_idx = list(df_ren.columns).index(name) + 1
            for row in range(2, ws.max_row + 1):
                c = ws.cell(row=row, column=col_idx)
                if isinstance(c.value, str):
                    s = c.value.replace(" ", "").replace(".", "").replace(",", ".")
                    try:
                        c.value = float(s)
                    except Exception:
                        pass
                c.number_format = "#.##0,00"

    bio.seek(0)
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=fatture.xlsx"},
    )


@app.get("/quotes/export")
async def export_quotes(format: str = "csv"):
    conn = get_connection()
    try:
        cur = conn.cursor()
        cur.execute(
            """
            SELECT customer, valid_until, total_amount, items_json, confidence
            FROM quotes
            ORDER BY created_at DESC
            """
        )
        rows = cur.fetchall()
    finally:
        conn.close()

    headers_it = ["Cliente", "Validità", "Totale", "Voci", "Confidenza"]

    if format.lower() in ("csv", "csv_it", "csv-it"):
        sio = StringIO()
        w = csv.writer(sio, delimiter=";")
        w.writerow(headers_it)
        for r in rows:
            w.writerow([
                r[0] or "",
                _fmt_date_it(r[1]),
                _fmt_num_it(r[2]),
                (r[3] or "").replace("\n", " | "),
                _fmt_num_it(r[4]),
            ])
        sio.seek(0)
        return StreamingResponse(
            iter([sio.getvalue()]),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": "attachment; filename=preventivi.csv"},
        )

    df = pd.DataFrame(rows, columns=["customer", "valid_until", "total_amount", "items_json", "confidence"])
    df["valid_until"] = pd.to_datetime(df["valid_until"], errors="coerce")

    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        sheet = "Preventivi"
        df_ita = df.rename(
            columns={
                "customer": "Cliente",
                "valid_until": "Validità",
                "total_amount": "Totale",
                "items_json": "Voci",
                "confidence": "Confidenza",
            }
        )
        df_ita.to_excel(writer, index=False, sheet_name=sheet)

        ws = writer.sheets[sheet]
        widths = {"Cliente": 30, "Validità": 12, "Totale": 14, "Voci": 50, "Confidenza": 12}
        for i, name in enumerate(["Cliente", "Validità", "Totale", "Voci", "Confidenza"], start=1):
            ws.column_dimensions[get_column_letter(i)].width = widths.get(name, 16)

        # Date
        date_col = 2
        for row in range(2, ws.max_row + 1):
            c = ws.cell(row=row, column=date_col)
            v = c.value
            if isinstance(v, str):
                try:
                    c.value = datetime.fromisoformat(v[:10])
                except Exception:
                    try:
                        parsed = pd.to_datetime(v, errors="coerce")
                        c.value = parsed.to_pydatetime() if pd.notna(parsed) else None
                    except Exception:
                        c.value = None
            c.number_format = "dd/mm/yyyy"

        # Numeri
        for idx in [3, 5]:
            for row in range(2, ws.max_row + 1):
                c = ws.cell(row=row, column=idx)
                if isinstance(c.value, str):
                    s = c.value.replace(" ", "").replace(".", "").replace(",", ".")
                    try:
                        c.value = float(s)
                    except Exception:
                        pass
                c.number_format = "#.##0,00"

    bio.seek(0)
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=preventivi.xlsx"},
    )
